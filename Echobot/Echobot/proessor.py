import openpyxl as xl
path1= "C:/Users/vishnu/Desktop/New.xlsx"
path2 = "C:/Users/vishnu/Desktop/Out.xlsx"
wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]
wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)
for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value
wb2.save(path2)